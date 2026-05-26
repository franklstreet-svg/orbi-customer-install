"""
catalog module — file-drop product catalog indexer.

The owner drops a CSV, Excel (.xlsx), or TSV file into Orby/Catalog/ —
exported from their POS (Square, Clover, QuickBooks), spreadsheet, or
typed by hand. This module:

  1. Watches the Catalog/ folder for new or changed files.
  2. Auto-detects which columns are SKU, name, price, stock, category,
     description. Owner doesn't have to map anything.
  3. Builds a fast keyword + numeric-SKU search index.
  4. Keeps the last 7 imports as snapshots in Catalog/.history/ so the
     owner can roll back a bad import.
  5. Exposes search() for the /chat endpoint to call BEFORE the LLM —
     so "do you carry a 3/4 inch brass elbow?" gets answered from real
     inventory data, never from an LLM guess.

Storage layout on the customer's machine:
  Orby/Catalog/
    inventory.csv              ← whatever the owner dropped in (kept)
    .index/                    ← parsed index (rebuilt as needed)
      items.json
      meta.json
    .history/                  ← last 7 imports for rollback
      2026-05-26T12-04-09__inventory.csv
      ...

Privacy: every byte stays inside the customer's Orby/ folder. The
catalog never moves to Frank's brain machine or HF Inference — only
the search RESULTS go into the LLM context when a chat answer needs
them, and only the minimal fields (name, price, stock) for the few
items matched.
"""

from __future__ import annotations

import csv
import json
import logging
import re
import shutil
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)

_LOCK = threading.Lock()
_WATCHER_THREAD: threading.Thread | None = None
_WATCHER_STOP = threading.Event()

# Column-detection hints. The owner's file might use ANY of these headers
# (case-insensitive). We score each header against these lists and assign
# the best match to its canonical role.
_COL_HINTS = {
    "sku":         ["sku", "part #", "part#", "part number", "part no", "item #", "item#",
                    "item number", "item no", "code", "product code", "product id",
                    "id", "model", "model #", "model number", "upc", "barcode"],
    "name":        ["name", "product", "product name", "item", "item name", "title",
                    "description short", "short description", "short desc"],
    "description": ["description", "desc", "details", "long description", "long desc",
                    "notes", "info"],
    "price":       ["price", "cost", "retail", "retail price", "msrp", "sell price",
                    "selling price", "amount", "unit price", "list price"],
    "stock":       ["stock", "qty", "quantity", "on hand", "in stock", "inventory",
                    "count", "available", "qty on hand", "qoh"],
    "category":    ["category", "cat", "type", "group", "department", "dept",
                    "class", "section"],
    "brand":       ["brand", "manufacturer", "mfg", "mfr", "make", "vendor"],
}

# Strip common currency prefixes/suffixes when parsing prices.
_PRICE_RE = re.compile(r'[\$£€¥]|\s|usd|cad|gbp|eur', re.IGNORECASE)

# Stop-words excluded from keyword tokens. Keep short; we want most words
# to count (a 3/4 elbow query has "3/4" + "elbow" — losing either kills
# the match).
_STOP_WORDS = {
    "the", "and", "or", "of", "for", "a", "an", "to", "in", "on", "at",
    "with", "by", "from",
}


# ── Public API ──────────────────────────────────────────────────────────


def _catalog_dir(data_dir: Path) -> Path:
    return data_dir / "Catalog"


def _index_dir(data_dir: Path) -> Path:
    return _catalog_dir(data_dir) / ".index"


def _history_dir(data_dir: Path) -> Path:
    return _catalog_dir(data_dir) / ".history"


def ensure_dirs(data_dir: Path) -> None:
    """Create Catalog/ + .index/ + .history/ so the owner can drop files in."""
    _catalog_dir(data_dir).mkdir(parents=True, exist_ok=True)
    _index_dir(data_dir).mkdir(parents=True, exist_ok=True)
    _history_dir(data_dir).mkdir(parents=True, exist_ok=True)


def status(data_dir: Path) -> dict:
    """Return what the owner needs to see in their dashboard:
    last imported file, item count, when, source path."""
    meta_path = _index_dir(data_dir) / "meta.json"
    if not meta_path.exists():
        return {
            "indexed": False,
            "item_count": 0,
            "last_indexed_at": None,
            "source_file": None,
            "columns_detected": {},
        }
    try:
        return json.loads(meta_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"indexed": False, "item_count": 0}


def search(data_dir: Path, query: str, limit: int = 10) -> list[dict]:
    """Find items matching the query. Returns a list of item dicts with
    fields {sku, name, description, price, stock, category, brand, score}.

    Ranking:
      - Exact SKU match scores 1000 (always wins; SKUs are unique)
      - Token overlap with name scores 10 per matched token
      - Token overlap with description / brand / category scores 3 per token
      - Numeric / dimensional tokens (e.g. "3/4", "12mm") score 20 per token
        because they're the most-discriminating part of a parts query
    Returns top `limit` items sorted by score, score >= 1.
    """
    items_path = _index_dir(data_dir) / "items.json"
    if not items_path.exists():
        return []
    try:
        items = json.loads(items_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []

    q_lower = query.lower()
    q_tokens = _tokenize(query)
    if not q_tokens:
        return []

    # Exact SKU match short-circuit — if any token matches an SKU exactly,
    # surface that item first.
    sku_hit: dict | None = None
    for item in items:
        sku_lower = (item.get("sku") or "").lower()
        if sku_lower and sku_lower in q_tokens:
            sku_hit = {**item, "score": 1000.0}
            break

    results = []
    for item in items:
        score = 0.0
        name_tokens = _tokenize(item.get("name", ""))
        desc_tokens = _tokenize(item.get("description", ""))
        meta_tokens = (
            _tokenize(item.get("brand", ""))
            | _tokenize(item.get("category", ""))
        )
        for tok in q_tokens:
            if tok in name_tokens:
                score += 20 if _looks_dimensional(tok) else 10
            elif tok in desc_tokens:
                score += 10 if _looks_dimensional(tok) else 3
            elif tok in meta_tokens:
                score += 3
        if score >= 1:
            results.append({**item, "score": round(score, 1)})

    results.sort(key=lambda r: r["score"], reverse=True)
    if sku_hit:
        results = [sku_hit] + [r for r in results if r.get("sku") != sku_hit.get("sku")]
    return results[:limit]


def reindex(data_dir: Path) -> dict:
    """Force a re-index of the newest catalog file. Returns the new status."""
    with _LOCK:
        return _do_reindex(data_dir)


def start_watcher(data_dir: Path, poll_seconds: int = 10) -> None:
    """Start the background folder watcher. Idempotent — safe to call twice."""
    global _WATCHER_THREAD
    if _WATCHER_THREAD and _WATCHER_THREAD.is_alive():
        return
    ensure_dirs(data_dir)
    _WATCHER_STOP.clear()
    _WATCHER_THREAD = threading.Thread(
        target=_watcher_loop,
        args=(data_dir, poll_seconds),
        daemon=True,
        name="catalog-watcher",
    )
    _WATCHER_THREAD.start()
    log.info("catalog watcher started (data_dir=%s, poll=%ss)", data_dir, poll_seconds)


def stop_watcher() -> None:
    _WATCHER_STOP.set()


# ── Internals ───────────────────────────────────────────────────────────


def _watcher_loop(data_dir: Path, poll_seconds: int) -> None:
    """Background loop: check the Catalog/ folder mtime every poll_seconds.
    If a file changed since last index, re-index. Crash-safe — any
    exception is logged and the loop continues."""
    last_source_mtime: float = -1.0
    while not _WATCHER_STOP.is_set():
        try:
            newest = _find_newest_source(data_dir)
            if newest is not None:
                mtime = newest.stat().st_mtime
                if mtime > last_source_mtime:
                    log.info("catalog watcher: %s changed, re-indexing", newest.name)
                    with _LOCK:
                        result = _do_reindex(data_dir, hint_path=newest)
                    if result.get("indexed"):
                        last_source_mtime = mtime
        except Exception as e:
            log.warning("catalog watcher loop error: %s", e)
        # Sleep in small chunks so stop signals fire quickly.
        for _ in range(poll_seconds):
            if _WATCHER_STOP.is_set():
                break
            time.sleep(1)


def _find_newest_source(data_dir: Path) -> Path | None:
    """Return the newest user-dropped catalog file, ignoring .index/.history."""
    cdir = _catalog_dir(data_dir)
    if not cdir.exists():
        return None
    candidates = [
        p for p in cdir.iterdir()
        if p.is_file()
        and p.suffix.lower() in {".csv", ".tsv", ".xlsx", ".xls", ".pdf", ".txt"}
        and not p.name.startswith(".")
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _do_reindex(data_dir: Path, hint_path: Path | None = None) -> dict:
    """Re-index the newest file in Catalog/. Snapshots the previous import
    into .history/ first. Returns the new status dict."""
    ensure_dirs(data_dir)
    src = hint_path or _find_newest_source(data_dir)
    if src is None:
        return {
            "indexed": False,
            "item_count": 0,
            "last_indexed_at": None,
            "source_file": None,
            "error": "no_catalog_file",
        }

    # Snapshot previous source into .history before overwriting the index.
    _snapshot_previous(data_dir)

    rows, raw_headers, parse_warning = _parse_file(src)
    if not rows:
        return {
            "indexed": False,
            "item_count": 0,
            "last_indexed_at": _now_iso(),
            "source_file": src.name,
            "error": parse_warning or "empty_file",
        }

    column_map = _detect_columns(raw_headers)
    items = _normalize_rows(rows, column_map)

    # Persist index + meta atomically.
    items_path = _index_dir(data_dir) / "items.json"
    meta_path = _index_dir(data_dir) / "meta.json"
    items_tmp = items_path.with_suffix(".json.tmp")
    items_tmp.write_text(json.dumps(items, ensure_ascii=False), encoding="utf-8")
    items_tmp.replace(items_path)

    meta = {
        "indexed": True,
        "item_count": len(items),
        "last_indexed_at": _now_iso(),
        "source_file": src.name,
        "source_path": str(src),
        "columns_detected": column_map,
        "raw_headers": raw_headers,
    }
    if parse_warning:
        meta["warning"] = parse_warning
    meta_tmp = meta_path.with_suffix(".json.tmp")
    meta_tmp.write_text(json.dumps(meta, indent=2), encoding="utf-8")
    meta_tmp.replace(meta_path)

    log.info(
        "catalog re-indexed: %d items from %s (cols: %s)",
        len(items), src.name,
        ", ".join(f"{k}={v}" for k, v in column_map.items() if v is not None),
    )
    return meta


def _snapshot_previous(data_dir: Path) -> None:
    """Copy the current source file into .history/ with a timestamp prefix,
    then prune to keep only the last 7."""
    meta_path = _index_dir(data_dir) / "meta.json"
    if not meta_path.exists():
        return
    try:
        prev = json.loads(meta_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return
    prev_source = prev.get("source_path")
    if not prev_source:
        return
    p = Path(prev_source)
    if not p.exists():
        return
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%S")
    dest = _history_dir(data_dir) / f"{stamp}__{p.name}"
    try:
        shutil.copy2(p, dest)
    except Exception as e:
        log.warning("could not snapshot %s: %s", p, e)
    # Prune: keep newest 7
    history = sorted(_history_dir(data_dir).glob("*"), key=lambda x: x.stat().st_mtime, reverse=True)
    for old in history[7:]:
        try:
            old.unlink()
        except Exception:
            pass


def _parse_file(path: Path) -> tuple[list[dict], list[str], str | None]:
    """Dispatch to the right parser based on extension. Returns (rows,
    raw_headers, warning_or_None). Rows are dicts keyed by the raw
    header strings."""
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            return _parse_csv(path, delimiter=",")
        if suffix == ".tsv":
            return _parse_csv(path, delimiter="\t")
        if suffix == ".txt":
            # Try CSV first (one item per line), then plain text fallback.
            rows, hdrs, warn = _parse_csv(path, delimiter=",")
            if rows:
                return rows, hdrs, warn
            return _parse_txt(path)
        if suffix in {".xlsx", ".xls"}:
            return _parse_xlsx(path)
        if suffix == ".pdf":
            return [], [], "pdf_parsing_not_yet_supported"
    except Exception as e:
        log.warning("catalog parse failed for %s: %s", path, e)
        return [], [], f"parse_error: {e.__class__.__name__}"
    return [], [], f"unsupported_extension: {suffix}"


def _parse_csv(path: Path, delimiter: str) -> tuple[list[dict], list[str], str | None]:
    """Parse a CSV/TSV. Tolerates BOM, blank lines, weird quote styles."""
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        try:
            reader = csv.DictReader(f, delimiter=delimiter)
            rows = []
            for r in reader:
                # Drop empty rows + rows where every value is empty/whitespace
                if r and any((v or "").strip() for v in r.values()):
                    rows.append(r)
            headers = reader.fieldnames or []
            return rows, list(headers), None
        except csv.Error as e:
            return [], [], f"csv_error: {e}"


def _parse_xlsx(path: Path) -> tuple[list[dict], list[str], str | None]:
    """Parse an Excel sheet. Reads the first non-empty sheet, uses row 1
    as headers."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], [], "openpyxl_not_installed"
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    for ws in wb.worksheets:
        if ws.max_row and ws.max_row > 1:
            rows_iter = ws.iter_rows(values_only=True)
            headers_raw = next(rows_iter, None)
            if not headers_raw:
                continue
            headers = [str(h or "").strip() for h in headers_raw]
            # Drop trailing empty header columns
            while headers and not headers[-1]:
                headers.pop()
            if not headers:
                continue
            rows: list[dict] = []
            for row in rows_iter:
                if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                    continue
                d = {}
                for i, h in enumerate(headers):
                    v = row[i] if i < len(row) else None
                    d[h] = "" if v is None else str(v)
                rows.append(d)
            if rows:
                return rows, headers, None
    return [], [], "xlsx_no_data_sheet"


def _parse_txt(path: Path) -> tuple[list[dict], list[str], str | None]:
    """Last-resort parser for plain text files — one product name per line.
    Treats the whole line as the 'name' column."""
    lines = [l.strip() for l in path.read_text(encoding="utf-8", errors="replace").splitlines()]
    lines = [l for l in lines if l]
    rows = [{"name": l} for l in lines]
    return rows, ["name"], None if rows else "empty_text_file"


def _detect_columns(headers: list[str]) -> dict[str, str | None]:
    """Map canonical role (sku/name/price/stock/category/description/brand)
    to the header in the file that best matches. Returns the best-match
    header string for each role, or None if no good match."""
    found: dict[str, str | None] = {role: None for role in _COL_HINTS}
    if not headers:
        return found
    norm = [(h, h.lower().strip()) for h in headers if h]
    # Two-pass: exact-match wins; then substring/contains fallback.
    for role, hints in _COL_HINTS.items():
        for h, h_norm in norm:
            if h_norm in hints:
                found[role] = h
                break
    for role, hints in _COL_HINTS.items():
        if found[role]:
            continue
        for h, h_norm in norm:
            if any(hint in h_norm for hint in hints):
                found[role] = h
                break
    # Special case: if no "name" detected but headers contain something
    # plausibly name-like, fall back to the first non-numeric column.
    if not found["name"]:
        for h, h_norm in norm:
            if h not in found.values():
                found["name"] = h
                break
    return found


def _normalize_rows(rows: list[dict], column_map: dict[str, str | None]) -> list[dict]:
    """Project raw rows into canonical item dicts."""
    items: list[dict] = []
    for r in rows:
        item = {
            "sku":         _val(r, column_map.get("sku")),
            "name":        _val(r, column_map.get("name")),
            "description": _val(r, column_map.get("description")),
            "price":       _parse_price(_val(r, column_map.get("price"))),
            "stock":       _parse_int(_val(r, column_map.get("stock"))),
            "category":    _val(r, column_map.get("category")),
            "brand":       _val(r, column_map.get("brand")),
        }
        # Skip items with neither sku nor name — nothing to search on.
        if not item["sku"] and not item["name"]:
            continue
        items.append(item)
    return items


def _val(row: dict, key: str | None) -> str:
    if not key:
        return ""
    return str(row.get(key, "") or "").strip()


def _parse_price(s: str) -> float | None:
    if not s:
        return None
    cleaned = _PRICE_RE.sub("", s).replace(",", "")
    if not cleaned:
        return None
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return None


def _parse_int(s: str) -> int | None:
    if not s:
        return None
    cleaned = re.sub(r"[^\-\d]", "", s)
    if not cleaned or cleaned == "-":
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def _tokenize(s: str) -> set[str]:
    """Lowercase + word-split. Preserves dimensional fragments like
    "3/4" and "1/2" because those are critical for parts lookups."""
    if not s:
        return set()
    s = s.lower()
    # Replace common separators with space, but KEEP the slash inside
    # fractions like 3/4 (which is two digits around a slash).
    s = re.sub(r"[^\w/.-]", " ", s)
    raw = s.split()
    return {t for t in raw if len(t) >= 2 and t not in _STOP_WORDS}


def _looks_dimensional(token: str) -> bool:
    """Token looks like a part-specific dimension (3/4, 12mm, 1.5, 1-1/4)."""
    return bool(re.match(r"^[\d./\-]+(?:mm|cm|in|inch|ft|m|kg|lb|oz|g)?$", token))


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
